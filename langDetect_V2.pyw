
import openpyxl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.workbook.workbook import Workbook
import langdetect
import os
from tkinter import *

if __name__ == "__main__":

    root = Tk()
    root.title("Language detector")
    root.resizable(0,0)

    full_path = StringVar()
    text_range = StringVar()
    path = os.path.dirname(str(full_path)).strip('"')
    my_file_name = os.path.splitext(os.path.basename((str(full_path))))[0]

    langs = []
    string_lst = []
 

    path_label = Label(root, text= "Enter the full path for the file:")
    path_label.grid(row=0, column=0, sticky=W, padx=5, pady=5)

    path_entry = Entry(root, textvariable=full_path)
    path_entry.grid(row=0, column=1, padx=5)

    column_label = Label(root, text ="Enter the column for detection:")
    column_label.grid(row=1, column=0, sticky=W, padx=5, pady=5)

    column_entry = Entry(root, textvariable=full_path)
    column_entry.grid(row=1, column=1, padx=5)

    prep_button = Button(root, text="Prepare file", command=lambda:[detect_lang, create_subfile, populate_lang])
    prep_button.grid(row=2, column=0, pady=50)

    compile_button = Button(root, text="Compile")
    compile_button.grid(row=2, column=1)

    
        

    def detect_lang():
        my_book = openpyxl.load_workbook((str(full_path)).strip('"')) #Strip quotation marks
        my_book_ws = my_book.active
        full_path.get()
        text_range.get()

        #Detects the lang of every row in a given column, populates the lang list and print each lang to the row in the right
        for col in my_book_ws.iter_cols(min_col= column_index_from_string(text_range),  max_col= column_index_from_string(text_range)):
            for cell in col:
                lang = (langdetect.detect(cell.value))

                if cell.value != None and lang not in langs:
                    langs.append(lang)
                cell.offset(row= 0, column= 1).value = lang #Offset one column to the right to populate with the dectected lang
        my_book.save((str(full_path)).strip('"'))
            
    def create_subfile():
        #Creates one file per each lang code in langs[]
        for lang in langs:
            new_wb = Workbook()
            
            new_full_path = (path + "\\" + lang + "_" + my_file_name + ".xlsx")
            new_full_path = new_full_path.lstrip('"')
            new_wb.save(new_full_path)

    def populate_lang():
        my_book = openpyxl.load_workbook((str(full_path)).strip('"')) #Strip quotation marks
        my_book_ws = my_book.active

        for cell in my_book_ws[text_range:text_range]:
                string_lst.append(cell.value)
                print(string_lst)
        for lang in langs:
            i = 1
            lang_wb = openpyxl.load_workbook(path + "\\" + lang + "_" + my_file_name + ".xlsx")
            lang_ws = lang_wb.active
            for string in string_lst:
                if string != None and langdetect.detect(string) == lang:
                    lang_ws.cell(row= i, column= 1, value=string)
                i += 1
            lang_wb.save(path + "\\" + lang + "_" + my_file_name + ".xlsx")

    def compile_target():
        for lang in langs:
                i = 1
                lang_wb = openpyxl.load_workbook(path + "\\" + lang + "_" + my_file_name + ".xlsx")
                lang_ws = lang_wb.active
                for string in string_lst:
                    if string != None and langdetect.detect(string) == lang:
                        lang_ws.cell(row= i, column= 1, value=string)
                    i += 1
                lang_wb.save(path + "\\" + lang + "_" + my_file_name + ".xlsx")


    create_subfile()
    populate_lang()
    compile_target()

    root.mainloop()



        
    